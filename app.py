import fitz  # PyMuPDF
import os
import re
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText
import sys

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

class StdoutRedirector:
    def __init__(self, text_widget):
        self.text_widget = text_widget

    def write(self, message):
        self.text_widget.configure(state='normal')
        self.text_widget.insert(tk.END, message)
        self.text_widget.see(tk.END)
        self.text_widget.configure(state='disabled')

    def flush(self):
        pass

def extract_text_from_pdfs(files):
    pdf_pages = []
    for pdf_path in files:
        filename = os.path.basename(pdf_path)
        try:
            doc = fitz.open(pdf_path)
            for page_num, page in enumerate(doc, start=1):
                page_text = page.get_text()
                pdf_pages.append((f"{os.path.splitext(filename)[0]}_Page_{page_num}", page_text))
                print(f"Extracted text from: {filename} - Page {page_num}")
            doc.close()
        except Exception as e:
            print(f"Failed to process {filename}: {e}")
    return pdf_pages
def parse_courses(lines):
    semesters, current_semester = {}, None
    sem_pat = re.compile(r"^(Fall|Spring|Summer|Winter)\s+\d{4} - .+")
    # Regular completed courses
    course_pat = re.compile(r"^([A-Z]{2,4})\s*(\d{3})\s+(.+?)\s+(\d+\.\d{2})\s+([A-F][+-]?)\s+(\d+\.\d{2})$")
    # In-progress courses
    in_progress_pat = re.compile(r"^([A-Z]{2,4})\s*(\d{3})\s+(.+?)\s+(\d+\.\d{2})\s+IN PROGRESS$")

    for line in lines:
        line = line.strip()
        if sem_pat.match(line):
            current_semester = line
            semesters.setdefault(current_semester, [])
        elif (m := course_pat.match(line)) and current_semester:
            semesters[current_semester].append({
                "course_code": m.group(1) + " " + m.group(2),
                "course_title": m.group(3).strip(),
                "credit_hours": float(m.group(4)),
                "letter_grade": m.group(5),
                "achieved_points": float(m.group(6))
            })
        elif (m := in_progress_pat.match(line)) and current_semester:
            semesters[current_semester].append({
                "course_code": m.group(1) + " " + m.group(2),
                "course_title": m.group(3).strip(),
                "credit_hours": float(m.group(4)),
                "letter_grade": "IN PROGRESS",
                "achieved_points": "IN PROGRESS"  # No points yet
            })
    return semesters


def parse_transcript_text(text):
    lines = text.splitlines()
    
    # Initialize default values
    name = "Unknown"
    student_id = "Unknown"
    
    # Search for name and ID in the first 20 lines
    for line in lines[2:20]:
        line = line.strip()
        if ',' in line and all(x.isalpha() or x in " ,-." for x in line.replace(',', '')) and name == "Unknown":
            name = line
        if re.match(r"^6\d{8}$", line) and student_id == "Unknown":
            student_id = line
        if name != "Unknown" and student_id != "Unknown":
            break

    # Extract major as before
    major = next((line.split(":")[1].strip() for line in lines if "Major :" in line), "Unknown")
    
    semester_courses = parse_courses(lines)

    semester_order = {"Spring": 1, "Summer": 2, "Fall": 3, "Winter": 4}
    sorted_semesters = sorted(semester_courses.items(), key=lambda x: (int(x[0].split()[1]), semester_order.get(x[0].split()[0], 0)))

    return {"name": name, "student_id": student_id, "major": major, "semesters": sorted_semesters}



def export_to_excel(transcript, output_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transcript"

    headers = ["NAME", "ID", "Major", "Year", "Term", "Course Code", "Course Credit", "Score"]
    ws.append(headers)

    for semester, courses in transcript['semesters']:
        year, term = semester.split()[:2]
        for c in courses:
            row = [
                transcript['name'],
                transcript['student_id'],
                transcript.get('major', 'Unknown'),
                year,
                term,
                c['course_code'],
                c.get('credit_hours', ''),
                c['letter_grade']
            ]
            ws.append(row)

    # Optionally, style header row bold
    for col_num in range(1, len(headers) + 1):
        ws[f"{get_column_letter(col_num)}1"].font = Font(bold=True)

    wb.save(output_file)
    print(f"Excel file saved as {output_file}")

def combine_all_excels(excel_files, output_folder, combined_filename="combined_transcripts.xlsx"):
    combined_wb = openpyxl.Workbook()
    combined_ws = combined_wb.active
    combined_ws.title = "Combined"

    headers_written = False
    seen = set()  # to avoid duplicates based on key

    for file in excel_files:
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            rows = list(ws.iter_rows(values_only=True))
            if not headers_written:
                combined_ws.append(rows[0])  # headers
                headers_written = True

            for row in rows[1:]:
                # Create a unique key to avoid duplicates, e.g. ID + Course Code + Year + Term
                key = (row[1], row[5], row[3], row[4])
                if key not in seen:
                    combined_ws.append(row)
                    seen.add(key)

        except Exception as e:
            print(f"Failed to read {os.path.basename(file)}: {e}")

    combined_path = os.path.join(output_folder, combined_filename)
    combined_wb.save(combined_path)
    print(f"Combined Excel saved as {combined_path}")


def process_files(files, output_folder):
    excel_files_created = []

    for pdf_path in files:
        filename = os.path.basename(pdf_path)
        base_filename = os.path.splitext(filename)[0]

        try:
            doc = fitz.open(pdf_path)
            all_courses = {}
            transcript_info = None

            for page_num, page in enumerate(doc, start=1):
                page_text = page.get_text()

                if page_num == 1:
                    transcript_info = parse_transcript_text(page_text)
                    all_courses = {
                        sem: courses for sem, courses in transcript_info['semesters']
                    }
                else:
                    semester_courses = parse_courses(page_text.splitlines())
                    for sem, courses in semester_courses.items():
                        all_courses.setdefault(sem, []).extend(courses)

            doc.close()

            if transcript_info:
                transcript_info['semesters'] = sorted(
                    all_courses.items(),
                    key=lambda x: (
                        int(x[0].split()[1]),
                        {"Spring": 1, "Summer": 2, "Fall": 3, "Winter": 4}
                        .get(x[0].split()[0], 0)
                    )
                )
                excel_filename = os.path.join(output_folder, f"{base_filename}.xlsx")
                export_to_excel(transcript_info, excel_filename)
                excel_files_created.append(excel_filename)  # Save path in order
            else:
                print(f"Failed to parse transcript from {filename}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to process {filename}:\n{e}")  # <-- popup here
            print(f"Failed to process {filename}: {e}")
    # Pass the excel files in the order created for combining
    combine_all_excels(excel_files_created, output_folder)

def select_files_and_process():
    files = filedialog.askopenfilenames(title="Select PDF files", filetypes=[("PDF Files", "*.pdf")])
    if files:
        process_files(files, output_folder_path.get())

def select_folder_and_process():
    folder = filedialog.askdirectory(title="Select PDF folder")
    if folder:
        files = [os.path.join(folder, f) for f in os.listdir(folder) if f.lower().endswith('.pdf')]
        if files:
            process_files(files, output_folder_path.get())
        else:
            messagebox.showwarning("No PDFs found", "No PDF files found in the selected folder.")

def browse_output_folder():
    folder = filedialog.askdirectory(title="Select Output Folder")
    if folder:
        output_folder_path.set(folder)
        lbl_output_folder.config(text=f"Output Folder: {folder}")
        btn_select_files.config(state=tk.NORMAL)
        btn_select_folder.config(state=tk.NORMAL)

def main():
    global txt_box, output_folder_path, lbl_output_folder, btn_select_files, btn_select_folder

    root = tk.Tk()
    root.title("Transcript Reader")
    root.geometry("850x650")
    root.resizable(False, False)
    icon_path = resource_path("icon.png")
    root.iconphoto(False, tk.PhotoImage(file=icon_path))

    # Minimalistic Colors
    BG_COLOR = "#FAFAFA"
    TXT_COLOR = "#222222"
    ACCENT_COLOR = "#000000"
    BTN_BG = "#E0E0E0"
    BTN_FG = "#222222"
    BTN_HOVER = "#C0C0C0"
    FRAME_BORDER_COLOR = "#DDDDDD"
    TXT_BG = "#FFFFFF"

    root.configure(bg=BG_COLOR)

    style = ttk.Style(root)
    style.theme_use('clam')

    style.configure('Accent.TButton',
                    background=BTN_BG,
                    foreground=BTN_FG,
                    font=('Segoe UI', 11),
                    padding=8,
                    borderwidth=0,
                    focusthickness=0)
    style.map('Accent.TButton',
              background=[('active', BTN_HOVER), ('pressed', BTN_HOVER)],
              relief=[('pressed', 'flat'), ('!pressed', 'flat')])

    style.configure("TLabel",
                    background=BG_COLOR,
                    foreground=TXT_COLOR,
                    font=("Segoe UI", 10))

    style.configure("Header.TLabel",
                    font=("Segoe UI", 22, "bold"),
                    background=BG_COLOR,
                    foreground=ACCENT_COLOR)

    output_folder_path = tk.StringVar()

    header_label = ttk.Label(root, text="Transcript Reader", style="Header.TLabel")
    header_label.pack(pady=(20, 15))

    top_frame = ttk.Frame(root, padding=5)
    top_frame.pack(fill=tk.X, padx=20, pady=(0, 15))
    top_frame.config(borderwidth=1, relief='solid')
    top_frame.configure(style='TopFrame.TFrame')
    style.configure('TopFrame.TFrame', background=BG_COLOR)

    btn_browse_output = ttk.Button(top_frame, text="Select Output Folder", command=browse_output_folder, style='Accent.TButton')
    btn_browse_output.pack(side=tk.LEFT)

    lbl_output_folder = ttk.Label(top_frame, text="Output Folder: (not selected)", anchor="w")
    lbl_output_folder.pack(side=tk.LEFT, padx=15, fill=tk.X, expand=True)

    btn_frame = ttk.Frame(root, padding=5)
    btn_frame.pack(pady=(0, 15), padx=20)
    btn_frame.config(borderwidth=1, relief='solid')
    btn_frame.configure(style='BtnFrame.TFrame')
    style.configure('BtnFrame.TFrame', background=BG_COLOR)

    btn_select_files = ttk.Button(btn_frame, text="Select Multiple PDFs", command=select_files_and_process, style='Accent.TButton', state=tk.DISABLED)
    btn_select_folder = ttk.Button(btn_frame, text="Select Folder", command=select_folder_and_process, style='Accent.TButton', state=tk.DISABLED)

    btn_select_files.grid(row=0, column=0, padx=5)
    btn_select_folder.grid(row=0, column=1, padx=5)

    txt_box = ScrolledText(root, width=100, height=30, font=("Consolas", 11), state='disabled',
                           bg=TXT_BG, fg=TXT_COLOR, relief="sunken", borderwidth=1)
    txt_box.tag_config('error', foreground='red')
    txt_box.pack(padx=20, pady=(0, 20), fill=tk.BOTH, expand=True)

    sys.stdout = StdoutRedirector(txt_box)
    sys.stderr = StdoutRedirector(txt_box)

    copyright_text = "© 2025 Jaden Peterson Wen"
    copyright_label = ttk.Label(root, text=copyright_text, font=("Segoe UI", 8, "italic"),
                                foreground="#777777", background=BG_COLOR)
    copyright_label.place(relx=1.0, rely=1.0, anchor='se', x=-10, y=-3)

    root.mainloop()

if __name__ == "__main__":
    main()
